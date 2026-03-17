import pandas as pd
import openpyxl

FILE = "C:/Users/junio/app-palmares/PALMARES 2024-2025.xlsx"
SEP = "=" * 70

xl = pd.ExcelFile(FILE, engine="openpyxl")
print(SEP)
print("FILE:", FILE)
print(SEP)
print(f"\nSHEET NAMES ({len(xl.sheet_names)} total):")
for i, name in enumerate(xl.sheet_names, 1):
    print(f"  {i:>2}. {name}")

for sheet_name in xl.sheet_names:
    print(f"\n{SEP}")
    print(f'SHEET: "{sheet_name}"')
    print(SEP)

    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb[sheet_name]
    merged = ws.merged_cells.ranges
    print(f"\n  openpyxl dims -> max_row={ws.max_row}  max_col={ws.max_column}")
    if merged:
        print(f"  Merged cell ranges ({len(merged)}):")
        for m in list(merged)[:12]:
            print(f"    {m}")
        if len(merged) > 12:
            print(f"    ... and {len(merged)-12} more")
    else:
        print("  No merged cells.")
    print("\n  First 3 RAW rows (values only):")
    for ri, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
        print(f"    Row {ri}: {list(row)}")
    wb.close()

    df = pd.read_excel(FILE, sheet_name=sheet_name, header=0, engine="openpyxl")
    unnamed_ratio = sum(1 for c in df.columns if str(c).startswith("Unnamed")) / max(len(df.columns), 1)
    header_used = 0
    if unnamed_ratio > 0.5:
        df = pd.read_excel(FILE, sheet_name=sheet_name, header=1, engine="openpyxl")
        header_used = 1

    df_clean = df.dropna(how="all")
    print(f"\n  pandas header row : {header_used}")
    print(f"  Columns ({len(df.columns)}): {list(df.columns)}")
    print(f"  Total rows        : {len(df)}  |  Non-empty rows: {len(df_clean)}")
    print("\n  First 5 data rows:")
    print(df_clean.head(5).to_string(index=True, max_colwidth=40))

print(f"\n{SEP}\nDONE.\n{SEP}")
